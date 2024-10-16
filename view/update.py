import datetime
import pandas as pd
import streamlit as st
import openpyxl


#df = openpyxl.load_workbook('namelist.xlsx')
#sheet = df.active #Assuming the data is in the first sheet

# Define CRUD functions

#def add_record(id, country, company, name, designation, dietary, contact, address, vehicle, posting_date, status, golf, golf_handicap): 
#    max_row = sheet.max_row + 1
#    sheet['A' + str(max_row)] = id 
#    sheet['B' + str(max_row)] = country
#    sheet['C' + str(max_row)] = company
#    sheet['D' + str(max_row)] = name
#    sheet['E' + str(max_row)] = designation
#    sheet['F' + str(max_row)] = dietary
#    sheet['G' + str(max_row)] = contact
#    sheet['H' + str(max_row)] = address
#    sheet['I' + str(max_row)] = vehicle
#    sheet['J' + str(max_row)] = posting_date
#    sheet['K' + str(max_row)] = status
#    sheet['L' + str(max_row)] = golf
#    sheet['M' + str(max_row)] = golf_handicap
#    df.save('namelist.xlsx')

#def del_record(id):
#    row_to_delete = None
#    for row in sheet.iter_rows(values_only=True):
#        if row[0] == id:
#            row_to_delete = row
#            break
#    if row_to_delete:
#        sheet.delete_rows(row_to_delete[0],1)
#        df.save('namelist.xlsx')
#    else:
#        st.warning('Not found.')

#def update_record(id, country, company, name, designation, dietary, contact, address, vehicle, posting_date, status, golf, golf_handicap, depost_date):
#    for row in sheet.iter_rows(values_only=True):
#        if row[0] == id:
#            row[1] = country
#            row[2] = company
#            row[3] = name
#            row[4] = designation
#            row[5] = dietary
#            row[6] = contact
#            row[7] = address
#            row[8] = vehicle
#            row[9] = posting_date
#            row[10]= status
#            row[11]= golf
#            row[12]= golf_handicap
#            df.save('namelist.xlsx')
#            break
#    else:
#        st.warning("Not found.")

#def view_all_employees():
#    data = []
#    for row in sheet.iter_rows(values_only=True):
#        data.append(row)
#    return data

 # Download button
#if st.button("Download current list"):
#    temp_file = "temp_database.xlsx"
#    df.save(temp_file)
        
    # Download the file
#    with open(temp_file, "rb") as file:
#        st.download_button(label="Download", data=file, file_name="namelist.xlsx")


#st.header("What would you like to do?")
# View All Employees
#with st.expander("View all"):
#    namelist = view_all_employees()
#    if namelist:
#        df = pd.DataFrame(namelist, columns=["ID", "Country", "Company", "Name", "Designation", "Dietary Restriction", "Contact No.", "Address", "Vehicle No.", "Posting Date", "Status", "Golf", "Golf Handicap", "De-posted Date"])
#        st.dataframe(df.iloc[1:],hide_index=True)
#    else:
#        st.warning("No employees found.")

# Add Employee
#with st.expander("Add"):
#    country = st.text_input("Country")
#    company = st.text_input("Company")
#    name = st.text_input("Name")
#    designation = st.text_input("Designation")
#    dietary = st.text_input("Dietary Restriction")
#    contact = st.text_input("Contact No.")
#    address = st.text_input("Address")
#    vehicle = st.text_input("Vehicle No.")
#    posting_date = st.date_input("Posting Date")
#    status = "Active"
#    golf = st.radio("Plays golf?", ("Yes", "No"))
#    golf_handicap = st.text_input("Golf Handicap (Please key in N.A if no handicap): ")

#    if st.button("Add"):
#        add_record(id, country, company, name, designation, dietary, contact, address, vehicle, posting_date, status, golf, golf_handicap)
#        st.success("Added successfully!") 

# Delete Employee
#with st.expander("Delete"):
    #id = st.number_input("Employee ID")

    #if st.button("Delete"):
    #    del_record(id)
    #    st.success("Deleted successfully!")

# Update Employee
#with st.expander("Update"):
    #id = st.number_input("Current Employee ID")
    #country = st.text_input("New Country")
    #company = st.text_input("New Company")
    #name = st.text_input("New Name")
    #designation = st.text_input("New Designation")
    #dietary = st.text_input("New Dietary Restriction")
    #contact = st.text_input("New Contact No.")
    #address = st.text_input("New Address")
    #vehicle = st.text_input("New Vehicle No.")
    #posting_date = st.date_input("New Posting Date")
    #status = st.radio("New Status", ("Active", "De-posted"))
    #if status == "De-posted":
    #    depost_date = st.date_input("De-posted Date")
    #golf = st.radio("New Plays golf?", ("Yes", "No"))
    #golf_handicap = st.radio("New Golf Hanficap?", ("No", "Yes"))
    #if golf_handicap == "Yes":
    #    golf_handicap = st.text_input("What's the handicap?")

    #if st.button("Update"):
    #    update_record(id, country, company, name, designation, dietary, contact, address, vehicle, posting_date, status, golf, golf_handicap, depost_date)
    #    st.success("Updated successfully!")

# Add a selectbox for choosing the action
action = st.selectbox("Choose an action:", ["Add Entry", "Update Entry", "Delete Entry", "View All Entries"])

# Load the Excel dataset
df = pd.read_excel("namelist.xlsx")

with st.expander("View list"):
    st.dataframe(df.iloc[:, 1:],hide_index=False)
# Define functions for each action
def add_entry():
    st.title("Add New Entry")

    name = st.text_input("Name")
    designation = st.text_input("Designation")
    country = st.text_input("Country")
    vehicle = st.text_input("Vehicle")
    company = st.text_input("Company")
    posting_date = st.date_input("Posting Date")
    contact_no = st.text_input("Contact No.")
    golf = st.radio("Golf", ("No", "Yes"))
    if golf == "Yes":
        golf_handicap = st.radio("Golf Handicap", ("No","Yes"))
        if golf_handicap == "Yes":
            golf_handicap = st.text_input("Golf Handicap")
    status = st.selectbox("Status", ["Active", "Deposted"])
    if status == "Deposted":
        deposted_date = st.date_input("De-posted Date")

    if st.button("Add Entry"):
        new_entry = {
            "Name": name,
            "Designation": designation,
            "Country": country,
            "Vehicle": vehicle,
            "Company": company,
            "Posting date": posting_date,
            "Contact No.": contact_no,
            "Golf": golf,
            "Golf handicap": golf_handicap,
            "Status": status
        }
        df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)
        df.to_excel("namelist.xlsx", index=False)
        st.success("Entry added successfully!")

def update_entry():
    st.title("Update Entry")
    
    entry_id = st.number_input("Entry ID (Choose from the list above)")

    if entry_id in df.index:
        selected_entry = df.loc[entry_id]

        name = st.text_input("Name", value=selected_entry["Name"])
        designation = st.text_input("Designation", value=selected_entry["Designation"])
        country = st.text_input("Country", value=selected_entry["Country"])
        vehicle = st.text_input("Vehicle(s)", value=selected_entry["Vehicle"])
        company = st.text_input("Company", value=selected_entry["Company"])
        #posting_date = st.date_input("Posting Date", value=selected_entry["Posting Date"])
        contact_no = st.text_input("Contact No.", value=selected_entry["Contact No."])
        golf = st.text_input("Golf (Yes/No)", value=selected_entry["Golf"])
        golf_handicap = st.text_input("Golf Handicap", value=selected_entry["Golf Handicap"])
        #status = st.selectbox("Status", ["Active", "Deposted"], index=df.loc[entry_id]["Status"])

        if st.button("Update Entry"):
            df.loc[entry_id] = {
                "Name": name,
                "Designation": designation,
                "Country": country,
                "Vehicle(s)": vehicle,
                "Company": company,
                #"Posting date": posting_date,
                "Contact No.": contact_no,
                "Golf": golf,
                "Golf handicap": golf_handicap,
                #"Status": status
            }
            df.to_excel("namelist.xlsx", index=False)
            st.success("Entry updated successfully!")
    else:
        st.warning("Entry not found.")

def delete_entry():
    st.title("Delete Entry")
    entry_id = st.number_input("Entry ID (Choose from the list above)")

    if st.button("Delete Entry"):
        if entry_id in df.index:
            df = df.drop(entry_id, axis=0)
            df.to_excel("namelist.xlsx", index=False)
            st.success("Entry deleted successfully!")
        else:
            st.warning("Entry not found.")

def view_all_entries():
    st.title("View All Entries")

    #st.dataframe(df)
    st.dataframe(df.iloc[:,1:],hide_index=True)

# Run the selected action
if action == "Add Entry":
    add_entry()
elif action == "Update Entry":
    update_entry()
elif action == "Delete Entry":
    delete_entry()
else:
    view_all_entries()